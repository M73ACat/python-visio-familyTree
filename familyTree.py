'''
Author: M73ACat
Date: 2023-04-27 22:42:00
LastEditors: M73ACat (mengyao.wang0521@qq.com)
LastEditTime: 2023-05-02 22:19:16
Description: 
Copyright (c) 2023 by M73ACat, All Rights Reserved. 
'''

import win32com.client as win32
import os
import openpyxl

path = os.path.dirname(os.path.abspath(__file__))

def xlsx_load(workbook, sheetname):
    sheet = workbook[sheetname]
    msg_dic = {}
    for row in range(1, sheet.max_row):
        time = sheet.cell(row=row+1, column=1).value
        if time not in msg_dic.keys():
            msg_dic[time] = [[],[]]
        msg_dic[time][0].append(sheet.cell(row=row+1, column=2).value)
        msg_dic[time][1].append(sheet.cell(row=row+1, column=3).value)
    for i in msg_dic.keys():
        yield i, msg_dic[i][0], msg_dic[i][1]


def shape_maker(time, names, flag, count):
    """  
    time: 从excel中读取的时间 / time read from excel
    names: 从excel中读取的名字 / names read from excel
    flag: 0为硕士，1为博士 / 0 for master, 1 for doctor
    count: shape的数量，配合间距以调整位置 / the number of shape, adjust the position with the y_gap
    """
    # 不复制，直接修改原shape / modify the original shape without copy
    if count < 1:
        time_shape_temp = time_shape
        # 复制name_shape以匹配学生数量 / copy name_shape to match the number of students
        name_shape_temp_list = [name_shape] + [name_shape.Duplicate() for _ in range(len(names)-1)]
        border_shape_temp = border_shape
    else:
        time_shape_temp = time_shape.Duplicate()
        name_shape_temp_list = [name_shape.Duplicate() for _ in range(len(names))]
        border_shape_temp = border_shape.Duplicate()
    # 调整文字、位置等 / adjust the text, position, etc.
    time_shape_temp.Text = str(time) + '级'
    time_shape_temp.CellsU("PinX").FormulaU = x_time
    time_shape_temp.CellsU("PinY").FormulaU = (y_default - y_gap * count)
    border_shape_temp.CellsU("PinX").FormulaU = (x_border_doctor if flag else x_border_master)
    border_shape_temp.CellsU("PinY").FormulaU = (y_default - y_gap * count)
    pinx = border_shape_temp.CellsU("PinX").ResultIU
    width = border_shape_temp.CellsU("Width").ResultIU
    start_x = pinx - width/2
    # 多个名字在框中居中并等间距排列 / multiple names are centered in the box and arranged equidistantly
    name_pinx_list = [start_x + width/(len(names)+1)*(i+1) for i in range(len(names))]
    for num, name_shape_temp in enumerate(name_shape_temp_list):
        name_shape_temp.Text = names[num]
        name_shape_temp.CellsU("PinX").FormulaU = name_pinx_list[num]
        name_shape_temp.CellsU("PinY").FormulaU = border_shape_temp.CellsU("PinY").ResultIU
    if count > 0:
        # 创建并调整起始和终止点 / create and adjust the start and end points
        connect_line = conn_shape.Duplicate()
        connect_line.CellsU("BeginX").FormulaU = x_header
        connect_line.CellsU("EndX").FormulaU = border_shape_temp.CellsU("PinX").ResultIU
        connect_line.CellsU("BeginY").FormulaU = y_header_lower
        connect_line.CellsU("EndY").FormulaU = border_shape_temp.CellsU("PinY").ResultIU + border_shape_temp.CellsU("Height").ResultIU / 2
        # 调整连接线中间点的相对位置 / adjust the relative position of the middle point of the connection line
        connect_line.CellsU("Geometry1.Y2").FormulaU = '6.5 mm'
        connect_line.CellsU("Geometry1.Y3").FormulaU = '6.5 mm'

def main():
    # 读取excel / read excel
    workbook = openpyxl.load_workbook(path+'/信息.xlsx')
    worksheet = xlsx_load(workbook,'Sheet1')

    for num, msg in enumerate(worksheet):
        # 读取excel中的信息 / read information from excel
        time, names, student_type = msg
        # 博士和硕士分开 / separate the message of doctor and master
        doctor_list = [names[i] for i in range(len(names)) if '博士' in student_type[i]]
        master_list = [names[i] for i in range(len(names)) if '硕士' in student_type[i]]
        # 创建shape / create shape
        shape_maker(time, master_list, 0, num)
        if len(doctor_list):
            shape_maker(time, doctor_list, 1, num)

if __name__ == '__main__':

    visio = win32.gencache.EnsureDispatch("Visio.Application")
    # visio.Visible = 0
    vdoc = visio.Documents.Open("%s/source.vsdx"%path)

    page = vdoc.Pages.Item(1)

    # 用shape.NameID获取shape / Get the shape with shape.NameID
    header_shape = page.Shapes.ItemU('Sheet.10')
    time_shape = page.Shapes.ItemU('Sheet.11')
    border_shape = page.Shapes.ItemU('Sheet.18')
    name_shape = page.Shapes.ItemU('Sheet.19')
    conn_shape = page.Shapes.ItemU('Sheet.20')

    # 通过在page.Shapes中的顺序获取shape / Get the shape with the order in page.Shapes
    # header_shape = page.Shapes(1)
    # time_shape = page.Shapes(2)
    # border_shape = page.Shapes(3)
    # name_shape = page.Shapes(4)
    # conn_shape = page.Shapes(5)

    # 获取基本的位置信息 / Get the basic position information
    x_border_doctor = page.Shapes.ItemU('Sheet.14').CellsU("PinX").ResultIU
    x_border_master = page.Shapes.ItemU('Sheet.12').CellsU("PinX").ResultIU
    x_time = time_shape.CellsU("PinX").ResultIU
    x_header = header_shape.CellsU("PinX").ResultIU
    y_header_lower = header_shape.CellsU("PinY").ResultIU - header_shape.CellsU("Height").ResultIU / 2
    y_default = border_shape.CellsU("PinY").ResultIU
    y_gap = 40 / 25.4

    main()
    
    vdoc.SaveAs("%s/familyTree.vsdx"%path)